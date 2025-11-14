# st_app.py (VERSIÓN DEMO)
import io
import json
import os
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill

# Importar las clases/utilidades del gestor existente (mismo directorio)
from gestor_oop import GestorStock, norm_codigo, norm_talla, parse_fecha_excel

# -------------------
# Rutas DEMO por defecto (data_example/)
# -------------------
BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data_example"

DEFAULT_INV_PATH = str(DATA_DIR / "datos_almacen_example.json")
DEFAULT_PREV_PATH = str(DATA_DIR / "prevision_example.json")
DEFAULT_TALL_PATH = str(DATA_DIR / "talleres_example.json")
DEFAULT_CLI_PATH = str(DATA_DIR / "clientes_example.json")


# --------------
# Helpers básicos
# --------------
@st.cache_resource
def get_manager(
    path_inventario: str = DEFAULT_INV_PATH,
    path_prevision: str = DEFAULT_PREV_PATH,
    path_talleres: str = DEFAULT_TALL_PATH,
    path_clientes: str = DEFAULT_CLI_PATH,
) -> GestorStock:
    """
    Crea una única instancia por sesión de Streamlit.
    En la demo usa por defecto los JSON de data_example/.
    """
    mgr = GestorStock(
        path_inventario=path_inventario,
        path_prevision=path_prevision,
        path_talleres=path_talleres,
        path_clientes=path_clientes,
    )

    # Si el backend no define EXPORT_DIR, usamos carpeta local de exportación
    if not hasattr(mgr, "EXPORT_DIR"):
        export_dir = os.path.join(os.path.dirname(path_inventario), "EXPORTAR_CSV")
        mgr.EXPORT_DIR = export_dir

    return mgr


def _to_df(lista: List[Dict]) -> pd.DataFrame:
    if not lista:
        return pd.DataFrame()
    return pd.DataFrame(lista)


def _success(msg: str):
    st.success(msg, icon="✅")


def _warn(msg: str):
    st.warning(msg, icon="⚠️")


def _error(msg: str):
    st.error(msg, icon="❌")


def _info(msg: str):
    st.info(msg, icon="ℹ️")


def _to_df(lista: List[Dict]) -> pd.DataFrame:
    if not lista:
        return pd.DataFrame()
    return pd.DataFrame(lista)


def _success(msg: str):
    st.success(msg, icon="✅")


def _warn(msg: str):
    st.warning(msg, icon="⚠️")


def _error(msg: str):
    st.error(msg, icon="❌")


def _info(msg: str):
    st.info(msg, icon="ℹ️")


# ---- Export helpers ----
def _run_export_all(mgr: GestorStock):
    """Lanza la exportación completa a CSV en la ruta definida por el gestor."""
    try:
        if hasattr(mgr, "_exportar_todos_los_datos"):
            mgr._exportar_todos_los_datos()

            export_csv_dir = getattr(
                mgr, "EXPORT_CSV_DIR", getattr(mgr, "EXPORT_DIR", "(ruta no definida)")
            )
            _success(f"Exportación completa realizada en: {export_csv_dir}")
            set_last_update(mgr, "Exportación CSV (pack completo)")
        else:
            _error(
                "El backend no expone '_exportar_todos_los_datos'. Actualiza gestor_oop.py."
            )
    except Exception as e:
        _error(f"Error exportando CSV: {e}")


def _run_export_stock_negativo(mgr: GestorStock):
    try:
        if hasattr(mgr, "_exportar_stock_negativo"):
            mgr._exportar_stock_negativo()

            export_csv_dir = getattr(
                mgr, "EXPORT_CSV_DIR", getattr(mgr, "EXPORT_DIR", "(ruta no definida)")
            )
            _success(f"Exportado informe de stock negativo en: {export_csv_dir}")
        else:
            _error("El backend no expone '_exportar_stock_negativo'.")
    except Exception as e:
        _error(f"Error exportando stock negativo: {e}")


def _export_excel_pack(mgr: GestorStock):
    """
    1) Lanza el exportador de CSV del backend.
    2) Genera los 5 Excel IMPRIMIR_XX_... con formato para imprimir.
    """
    # Directorios: CSV de entrada y carpeta de salida para IMPRIMIR
    project_root = Path(__file__).resolve().parent.parent

    csv_dir = Path(getattr(mgr, "EXPORT_CSV_DIR", project_root / "exports" / "csv"))
    imprimir_dir = Path(
        getattr(mgr, "IMPRIMIR_DIR", project_root / "exports" / "excel" / "imprimir")
    )

    csv_dir.mkdir(parents=True, exist_ok=True)
    imprimir_dir.mkdir(parents=True, exist_ok=True)

    # 1) nos aseguramos de que los CSV están frescos
    if hasattr(mgr, "_exportar_todos_los_datos"):
        mgr._exportar_todos_los_datos()
    else:
        _error(
            "El backend no expone '_exportar_todos_los_datos'. Actualiza gestor_oop.py."
        )
        return

    hoy_str = datetime.now().strftime("%Y-%m-%d")

    config = {
        "00": ("stock_actual", ["STOCK", "stock"], "stock"),
        "03": ("pedidos_pendientes", None, None),
        "04": ("ordenes_fabricacion", None, None),
        "05": ("stock_estimado", ["STOCK_ESTIMADO", "stock_estimado"], "estimado"),
        "06": ("orden_corte_sugerida", None, None),
    }

    generados: list[str] = []

    for prefijo, (base_name, qty_candidates, tipo) in config.items():
        patrones = [
            f"{prefijo}_{base_name}_*.csv",
            f"*{prefijo}*{base_name}*.csv",
        ]
        csv_path = None
        for pat in patrones:
            matches = sorted(csv_dir.glob(pat))
            if matches:
                csv_path = matches[-1]
                break

        if not csv_path or not csv_path.exists():
            # si no se encuentra, pasamos a la siguiente hoja
            continue

        # lee CSV con separador ;
        df = pd.read_csv(csv_path, sep=";", dtype=str)
        xlsx_name = f"IMPRIMIR_{prefijo}_{base_name}_{hoy_str}.xlsx"
        xlsx_path = imprimir_dir / xlsx_name

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Hoja1")
            ws = writer.sheets["Hoja1"]

            # 1) Reglas base de colores
            if prefijo in ("00", "05"):
                _excel_color_stock_ranges(ws, df, qty_col_candidates=qty_candidates)
            elif prefijo == "03":
                _excel_color_pend_by_month(ws, df, date_col="FECHA")
            elif prefijo == "04":
                _excel_color_by_column_palette(ws, df, col="FECHA")
            elif prefijo == "06":
                _excel_color_by_column_palette(ws, df, col="MODELO")

            # 2) Totales por modelo (00,03,04,05)
            if prefijo in ("00", "03", "04", "05"):
                _excel_highlight_totals_by_talla(ws, df, talla_col="TALLA")

            # 3) Cabecera + total general
            highlight_last = prefijo in ("00", "03", "04", "05")
            _excel_yellow_header_and_total(ws, highlight_last=highlight_last)

            # 4) Añade borde fino a las celdas
            _excel_add_borders(ws)

        generados.append(str(xlsx_path))

    if generados:
        _success("Excel IMPRIMIR generados:\n" + "\n".join(f"- {p}" for p in generados))
        set_last_update(mgr, "Exportación Excel IMPRIMIR (pack)")
    else:
        _warn(
            "No se ha generado ningún Excel IMPRIMIR. Revisa que existan los CSV en la carpeta de exportación CSV y coincidan los patrones."
        )


def _timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


# --- Last Update (persistente en disco) ---
def _meta_path(mgr) -> str:
    base = os.path.dirname(mgr.ds_inventario.path)
    return os.path.join(base, "last_update.json")


def set_last_update(mgr, action: str, meta: dict | None = None):
    now = datetime.now()
    data = {
        "timestamp_iso": now.isoformat(timespec="seconds"),
        "timestamp_human": now.strftime("%d/%m/%Y %H:%M"),
        "action": action,
    }
    if meta:
        data.update(meta)
    try:
        with open(_meta_path(mgr), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    st.session_state["last_update"] = data
    return data


def get_last_update(mgr):
    if "last_update" in st.session_state:
        return st.session_state["last_update"]
    try:
        with open(_meta_path(mgr), "r", encoding="utf-8") as f:
            st.session_state["last_update"] = json.load(f)
            return st.session_state["last_update"]
    except Exception:
        return None


def _modo_dup_key(txt: str) -> str:
    return {
        "Descontar diferencia (recomendado)": "d",
        "Ignorar duplicadas": "i",
        "Procesar todo igualmente": "t",
    }[txt]


def _procesar_albaranes_df(df: pd.DataFrame, modo_txt: str, simular: bool):
    columnas = [
        "CodigoArticulo",
        "DesTalla",
        "Total",
        "SuPedido",
        "FechaAlbaran",
        "NumeroAlbaran",
    ]
    if not all(col in df.columns for col in columnas):
        _error(f"Faltan columnas necesarias: {columnas}")
        return

    # Ledger de salidas ya registradas
    ya_registrado = defaultdict(int)
    for s in mgr.inventory.historial_salidas:
        try:
            k = (
                str(s.get("modelo", "")).strip().upper(),
                norm_talla(s.get("talla", "")),
                norm_codigo(s.get("pedido", "")),
                norm_codigo(s.get("albaran", "")),
            )
            ya_registrado[k] += int(s.get("cantidad", 0) or 0)
        except Exception:
            continue

    # Preparar líneas
    lineas, duplicadas = [], []
    for _, fila in df.iterrows():
        modelo = str(fila["CodigoArticulo"]).strip().upper()
        talla = norm_talla(fila["DesTalla"])

        # --- conversiones seguras ---
        val_total = fila["Total"]
        cantidad_excel = (
            int(val_total)
            if not pd.isna(val_total) and str(val_total).strip() != ""
            else 0
        )

        val_pedido = fila["SuPedido"]
        pedido = norm_codigo("" if pd.isna(val_pedido) else val_pedido)

        val_albar = fila["NumeroAlbaran"]
        albaran = norm_codigo("" if pd.isna(val_albar) else val_albar)

        val_fecha = fila["FechaAlbaran"]
        fecha = parse_fecha_excel(None if pd.isna(val_fecha) else val_fecha)
        # ----------------------------

        k = (modelo, talla, pedido, albaran)
        qty_prev = ya_registrado.get(k, 0)
        lineas.append(
            {
                "modelo": modelo,
                "talla": talla,
                "pedido": pedido,
                "albaran": albaran,
                "fecha": fecha,
                "cantidad_excel": cantidad_excel,
                "ya_prev": qty_prev,
            }
        )

        if qty_prev > 0:
            duplicadas.append((k, cantidad_excel, qty_prev))

    modo = _modo_dup_key(modo_txt)
    nuevas_salidas = 0
    import_rows = []
    pedidos_servicios = []
    pedidos_antes = list(mgr.prevision.pedidos)

    for L in lineas:
        modelo = L["modelo"]
        talla = L["talla"]
        pedido = L["pedido"]
        albaran = L["albaran"]
        fecha = L["fecha"]
        qty_excel = int(L["cantidad_excel"])
        qty_prev = int(L["ya_prev"])

        # decidir cantidad a aplicar según modo
        if modo == "d" and qty_prev > 0:
            aplicar = max(qty_excel - qty_prev, 0)
        elif modo == "i" and qty_prev > 0:
            aplicar = 0
        else:
            aplicar = qty_excel

        if aplicar <= 0:
            continue

        if not simular:
            # cliente: intenta resolver como en CLI (pendiente, info_modelos, vacío)
            cliente = ""
            for p in mgr.prevision.pedidos:
                if (
                    str(p.get("modelo", "")).strip().upper() == modelo
                    and norm_talla(p.get("talla", "")) == talla
                    and p.get("pedido", "") == pedido
                ):
                    cliente = p.get("cliente", "") or ""
                    if cliente:
                        break
            if not cliente:
                cliente = (
                    mgr.prevision.info_modelos.get(modelo, {}).get("cliente", "") or ""
                )
            mgr.inventory.register_exit(
                modelo=modelo,
                talla=talla,
                cantidad=aplicar,
                cliente=cliente,
                pedido=pedido,
                albaran=albaran,
                fecha=fecha,
            )
        nuevas_salidas += aplicar

        import_rows.append(
            {
                "FECHA": fecha,
                "MODELO": modelo,
                "TALLA": talla,
                "CANTIDAD": aplicar,
                "PEDIDO": pedido,
                "ALBARAN": albaran,
                "CLIENTE": "",
            }
        )

    # detectar pedidos servidos (como en tu versión)
    pedidos_despues = list(mgr.prevision.pedidos)
    set_antes = {
        (p["modelo"], norm_talla(p["talla"]), p["pedido"]) for p in pedidos_antes
    }
    set_despues = {
        (p["modelo"], norm_talla(p["talla"]), p["pedido"]) for p in pedidos_despues
    }
    servidos = set_antes - set_despues
    for m, t, ped in servidos:
        pedidos_servicios.append(
            {
                "MODELO": m,
                "TALLA": t,
                "PEDIDO": ped,
                "CANTIDAD_ORIGINAL": "",
                "CANTIDAD_SERVIDA": "",
                "RESTANTE": "",
                "FECHA_ALBARAN": "",
                "NUMERO_ALBARAN": "",
            }
        )

    if not simular:
        _success(
            f"Importación completada: {nuevas_salidas} movimientos de albaranes procesados."
        )
        set_last_update(mgr, f"Importación albaranes: {nuevas_salidas} movimientos")
        st.rerun()
    else:
        _info(f"Simulación: se procesarían {nuevas_salidas} movimientos.")

    if import_rows:
        st.dataframe(pd.DataFrame(import_rows), use_container_width=True)


def _procesar_pedidos_df(df: pd.DataFrame, simular: bool):
    columnas = [
        "CodigoArticulo",
        "DesTalla",
        "UnidadesPendientes",
        "SuPedido",
        "FechaEntrega",
        "NumeroPedido",
    ]
    if not all(col in df.columns for col in columnas):
        _error(f"Faltan columnas necesarias: {columnas}")
        return

    ya = {
        (
            str(p.get("modelo", "")).strip().upper(),
            norm_talla(p.get("talla", "")),
            p.get("pedido", ""),
        )
        for p in mgr.prevision.pedidos
    }
    nuevos, duplicados = 0, 0
    import_rows = []

    for _, fila in df.iterrows():
        modelo = str(fila["CodigoArticulo"]).strip().upper()
        talla = norm_talla(fila["DesTalla"])
        val = fila["UnidadesPendientes"]
        cantidad = int(val) if not pd.isna(val) else 0
        pedido = norm_codigo(fila["SuPedido"])
        numero_pedido = norm_codigo(fila["NumeroPedido"])
        fecha = parse_fecha_excel(fila["FechaEntrega"])
        cliente_resuelto = (
            mgr.inventory.info_modelos.get(modelo, {}).get("cliente", "") or ""
        )

        k = (modelo, talla, pedido)
        if k in ya:
            duplicados += 1
            continue

        if not simular:
            mgr.prevision.register_pending(
                modelo=modelo,
                talla=talla,
                cantidad=int(cantidad),
                pedido=pedido,
                cliente=cliente_resuelto,
                fecha=fecha or None,
                numero_pedido=numero_pedido or None,
            )
        nuevos += 1
        import_rows.append(
            {
                "FECHA": fecha,
                "PEDIDO": pedido,
                "NUMERO_PEDIDO": numero_pedido,
                "MODELO": modelo,
                "TALLA": talla,
                "CANTIDAD": int(cantidad),
                "CLIENTE": cliente_resuelto,
            }
        )

    if not simular:
        _success(
            f"Importación completada: {nuevos} nuevos pedidos añadidos. Ignorados duplicados: {duplicados}."
        )
        set_last_update(
            mgr, f"Importación pedidos: {nuevos} nuevos, {duplicados} duplicados"
        )
        st.rerun()
    else:
        _info(
            f"Simulación: se añadirían {nuevos} pedidos nuevos. Ignorados duplicados: {duplicados}."
        )

    if import_rows:
        st.dataframe(pd.DataFrame(import_rows), use_container_width=True)


def _modelo_labels_y_map(mgr):
    """
    Devuelve:
    - labels: lista de strings para el select ["M123 | Zapatilla - ROJO", ...]
    - label2model: dict {label -> "M123"}
    Usa info_modelos y, si está vacío, cae a los modelos presentes en el almacén.
    """
    info = mgr.inventory.info_modelos
    universe = sorted(set(list(info.keys()) + list(mgr.inventory.almacen.keys())))
    labels = []
    label2model = {}
    for m in universe:
        meta = info.get(m, {}) or {}
        desc = meta.get("descripcion", "")
        color = meta.get("color", "")
        extra = " - ".join([x for x in (desc, color) if x])
        label = f"{m} | {extra}" if extra else m
        labels.append(label)
        label2model[label] = m
    return labels, label2model


def _stock_actual(mgr, modelo: str, talla) -> int:
    """Devuelve el stock actual del modelo/talla con búsqueda robusta."""
    if not modelo or not talla:
        return 0
    stock_dict = mgr.inventory.almacen.get(modelo, {}) or {}

    # 1) clave normalizada (string tipo "34", "T.U.", etc.)
    key_norm = norm_talla(talla)
    if key_norm in stock_dict:
        try:
            return int(stock_dict[key_norm])
        except Exception:
            pass

    # 2) entero puro (por si las tallas se guardaron como int)
    try:
        t_int = int(str(talla).strip())
        if t_int in stock_dict:
            return int(stock_dict[t_int])
    except Exception:
        pass

    # 3) tal cual string sin normalizar
    t_str = str(talla)
    if t_str in stock_dict:
        try:
            return int(stock_dict[t_str])
        except Exception:
            pass

    return 0


def _fmt_pending_label(p: dict) -> str:
    """Devuelve una etiqueta legible para un pendiente ya preparado en pend_rows."""
    return (
        f"{p['IDX']} | {p['MODELO']} | T:{p.get('TALLA','')} | Q:{p.get('CANTIDAD',0)} | "
        f"Ped:{p.get('PEDIDO','-')} | Nº:{p.get('NUMERO_PEDIDO','-')} | "
        f"{p.get('CLIENTE','-')} | {p.get('FECHA','-')}"
    )


def parse_index_selection(s: str, max_idx: int) -> List[int]:
    """Convierte '1,3,5-8' en una lista de índices válidos (1..max_idx)."""
    sel = set()
    for token in s.replace(" ", "").split(","):
        if not token:
            continue
        if "-" in token:
            a, b = token.split("-", 1)
            if a.isdigit() and b.isdigit():
                a, b = int(a), int(b)
                if a <= b:
                    for x in range(a, b + 1):
                        if 1 <= x <= max_idx:
                            sel.add(x)
        else:
            if token.isdigit():
                x = int(token)
                if 1 <= x <= max_idx:
                    sel.add(x)
    return sorted(sel)


def _tallas_disponibles(mgr: GestorStock, modelo: str) -> List[str]:
    """
    Devuelve las tallas conocidas para un modelo, recopiladas de:
    - Stock (almacén)
    - Pedidos pendientes
    - Órdenes de fabricación
    """
    modelo = (modelo or "").upper().strip()
    tallas = set()

    # Stock
    if modelo and modelo in mgr.inventory.almacen:
        tallas.update(mgr.inventory.almacen.get(modelo, {}).keys())

    # Pendientes
    try:
        for _, p in mgr.prevision.list_pendings():
            if str(p.get("modelo", "")).upper().strip() == modelo:
                t = norm_talla(p.get("talla", ""))
                if t:
                    tallas.add(t)
    except Exception:
        pass

    # Fabricación
    try:
        for _, f in mgr.prevision.list_fabrication():
            if str(f.get("modelo", "")).upper().strip() == modelo:
                t = norm_talla(f.get("talla", ""))
                if t:
                    tallas.add(t)
    except Exception:
        pass

    # Limpieza y orden
    tallas = {norm_talla(t) for t in tallas if str(t).strip()}
    return sorted(tallas)


def talla_select(
    label: str, modelo: str, key_sel: str, key_txt: str, allow_manual: bool = True
) -> str:
    """
    Widget combinado:
    - Si hay tallas conocidas del modelo -> selectbox con opción de 'escribir manual'
    - Si no hay tallas -> text_input directo
    Devuelve SIEMPRE la talla elegida/escrita (string).
    """
    opciones = _tallas_disponibles(mgr, modelo)
    if opciones:
        opts = ["(elige)"] + opciones + (["(escribir manual)"] if allow_manual else [])
        choice = st.selectbox(label, options=opts, key=key_sel)
        if allow_manual and choice in ("(elige)", "(escribir manual)"):
            return st.text_input("Talla", key=key_txt).strip()
        return choice
    # Sin opciones conocidas: campo libre
    return st.text_input(label, key=key_txt).strip()


def _fix_bad_stock_values(mgr: GestorStock) -> Tuple[int, List[Dict]]:
    """
    Normaliza valores de stock no enteros o inválidos y además sanea
    tallas anómalas (None/NaN/""/"NAN"/"NA"): si traen cantidad != 0, la pone a 0.
    Devuelve (n_cambios, log_cambios).
    """
    import math

    cambios = []
    for modelo, tallas in list(mgr.inventory.almacen.items()):
        for talla, val in list(tallas.items()):
            original_val = val
            original_talla = talla
            # --- 1) normaliza valor a int (incluye strings "nan"/"none"/"")
            try:
                if val is None:
                    nuevo_val = 0
                elif isinstance(val, float) and math.isnan(val):
                    nuevo_val = 0
                elif isinstance(val, int):
                    nuevo_val = int(val)
                elif isinstance(val, float):
                    nuevo_val = int(val)
                elif isinstance(val, str):
                    s = val.strip().lower()
                    if s in ("nan", "none", ""):
                        nuevo_val = 0
                    else:
                        s = s.replace(",", ".")
                        nuevo_val = int(float(s))
                else:
                    nuevo_val = 0
            except Exception:
                nuevo_val = 0

            # --- 2) si la clave de talla es anómala, fuerza valor 0
            talla_str = str(talla).strip().upper()
            bad_key = (
                talla is None
                or (isinstance(talla, float) and math.isnan(talla))
                or talla_str in ("", "NAN", "NA", "NULL")
            )
            motivo = "VALOR_INVALIDO"
            if bad_key and nuevo_val != 0:
                nuevo_val = 0
                motivo = "TALLA_ANOMALA->VALOR_0"

            if nuevo_val != original_val:
                mgr.inventory.almacen[modelo][talla] = nuevo_val
                cambios.append(
                    {
                        "MODELO": modelo,
                        "TALLA": original_talla,
                        "ANTES": original_val,
                        "AJUSTADO_A": nuevo_val,
                        "MOTIVO": motivo,
                    }
                )

    if cambios:
        mgr.inventory.save()
    return len(cambios), cambios


def _fix_negativos_a_cero_gui(mgr: GestorStock) -> Tuple[int, str, List[Dict]]:
    """
    Ajusta cualquier stock < 0 a 0, guarda y crea un CSV de log en EXPORT_DIR.
    Devuelve (n_cambios, ruta_log, log_rows).
    """
    cambios = []
    for modelo, tallas in list(mgr.inventory.almacen.items()):
        for talla, val in list(tallas.items()):
            # fuerza a int de forma robusta
            try:
                v = int(val)
            except Exception:
                try:
                    v = int(float(str(val).replace(",", ".")))
                except Exception:
                    v = 0
            if v < 0:
                cambios.append(
                    {"MODELO": modelo, "TALLA": talla, "ANTES": v, "AJUSTADO_A": 0}
                )
                mgr.inventory.almacen[modelo][talla] = 0

    ruta_log = ""
    if cambios:
        mgr.inventory.save()

        # Usamos siempre la nueva carpeta de exportación CSV
        export_dir = getattr(mgr, "EXPORT_CSV_DIR", None)

        if export_dir is None:
            # Fallback ultra-seguro: carpeta exports/csv/ en raíz del proyecto
            project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            export_dir = os.path.join(project_root, "exports", "csv")

        os.makedirs(export_dir, exist_ok=True)

        ruta_log = os.path.join(export_dir, f"ajuste_negativos_{_timestamp()}.csv")

        pd.DataFrame(cambios).to_csv(ruta_log, index=False, encoding="utf-8-sig")

    return len(cambios), ruta_log, cambios


def _purge_bad_talla_keys_gui(
    mgr: GestorStock, only_zero: bool = True
) -> Tuple[int, str, List[Dict]]:
    """
    Elimina del JSON las entradas con claves de talla anómalas:
    None, NaN, "", "NAN", "NA", "NULL".
    Por defecto solo elimina si el valor de stock == 0 (only_zero=True).
    Devuelve (n_borradas, ruta_log, log_rows).
    """
    import math

    bad_labels = {"", "NAN", "NA", "NULL"}
    borradas = []
    for modelo, tallas in list(mgr.inventory.almacen.items()):
        # trabajamos sobre copia de claves para poder borrar
        for talla in list(tallas.keys()):
            talla_str = "" if talla is None else str(talla).strip().upper()
            is_bad = (
                talla is None
                or (isinstance(talla, float) and math.isnan(talla))
                or talla_str in bad_labels
            )
            if not is_bad:
                continue

            # valor actual
            val = tallas.get(talla, 0)
            try:
                v = int(val)
            except Exception:
                try:
                    v = int(float(str(val).replace(",", ".")))
                except Exception:
                    v = 0

            if (only_zero and v == 0) or (not only_zero):
                borradas.append({"MODELO": modelo, "TALLA": talla, "VALOR": v})
                del mgr.inventory.almacen[modelo][talla]

    ruta_log = ""
    if borradas:
        mgr.inventory.save()

        # Usamos siempre la carpeta nueva /exports/csv/
        export_dir = getattr(mgr, "EXPORT_CSV_DIR", None)

        if export_dir is None:
            # Fallback seguro por si el manager no lo tiene (demo o error)
            project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            export_dir = os.path.join(project_root, "exports", "csv")

        os.makedirs(export_dir, exist_ok=True)

        ruta_log = os.path.join(export_dir, f"purga_tallas_anomalas_{_timestamp()}.csv")

        pd.DataFrame(borradas).to_csv(ruta_log, index=False, encoding="utf-8-sig")

    return len(borradas), ruta_log, borradas


# ---- Styler helpers (colores) ----
from datetime import date


def _row_bg(row, color: str):
    # Devuelve un estilo por COLUMNA (lo que espera Styler.apply(axis=1))
    return pd.Series([f"background-color: {color}"] * len(row), index=row.index)


# Paletas para degradados de meses (pasado ↔ futuro)
_PAST_GREENS = [
    "#e8f5e9",  # -1 mes
    "#c8e6c9",
    "#a5d6a7",
    "#81c784",
    "#66bb6a",
    "#4caf50",
    "#388e3c",  # muy antiguo
]

_FUTURE_REDS = [
    "#ffebee",  # +1 mes
    "#ffcdd2",
    "#ef9a9a",
    "#e57373",
    "#ef5350",
    "#f44336",
    "#c62828",  # muy a futuro
]

BRIGHT_YELLOW = "FFFF00"  # amarillo Excel fuerte para cabeceras y totales
PASTEL_YELLOW = "FFF9C4"  # amarillo suave para semáforo


def _month_index(d: date) -> int:
    return d.year * 12 + d.month


def _month_delta_color(delta: int) -> str:
    """
    delta = idx_fecha - idx_hoy
    <0 => pasado (verde, más oscuro cuanto más antiguo)
    =0 => mes actual (sin color)
    >0 => futuro (rojo, más oscuro cuanto más lejano)
    """
    if delta == 0:
        return ""

    if delta < 0:
        idx = min(abs(delta) - 1, len(_PAST_GREENS) - 1)
        return _PAST_GREENS[idx]
    else:
        idx = min(delta - 1, len(_FUTURE_REDS) - 1)
        return _FUTURE_REDS[idx]


def _parse_date_flexible(val):
    """
    Acepta date/datetime o string en formatos típicos de tus CSV/Excel/JSON.
    """
    if isinstance(val, date):
        return val
    if not val:
        return None
    s = str(val).strip()
    # recorta por si viene con hora
    s10 = s[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s10, fmt).date()
        except Exception:
            pass
    return None


def style_stock_ranges(df, qty_col: str):
    """
    Pinta filas según qty_col:
      <=0 rojo, 0-10 naranja, 10-25 amarillo, >25 sin color.
    """

    def _styler(row):
        try:
            q = float(row.get(qty_col, 0) or 0)
        except Exception:
            q = 0
        if q <= 0:
            return _row_bg(row, "#ffcccc")  # rojo suave
        elif 0 < q <= 10:
            return _row_bg(row, "#ffe1b2")  # naranja suave
        elif 10 < q <= 25:
            return _row_bg(row, "#fff6b2")  # amarillo suave
        else:
            return pd.Series([""] * len(row), index=row.index)

    return df.style.apply(lambda r: _styler(r), axis=1)


def style_pend_by_month(df, date_col: str):
    """
    Pinta filas según mes de FECHA vs hoy, con degradado:
      - Pasado: verde, más oscuro cuanto más antiguo.
      - Mes actual: sin color.
      - Futuro: rojo, más oscuro cuanto más lejano.
    """
    today = date.today()
    cur_idx = _month_index(today)

    def _styler(row):
        d = _parse_date_flexible(row.get(date_col, ""))
        if not d:
            return pd.Series([""] * len(row), index=row.index)
        idx = _month_index(d)
        delta = idx - cur_idx
        color = _month_delta_color(delta)
        if not color:
            return pd.Series([""] * len(row), index=row.index)
        return _row_bg(row, color)

    return df.style.apply(lambda r: _styler(r), axis=1)


def style_by_column_palette(df: pd.DataFrame, col: str):
    """
    Asigna un color de fondo por cada valor distinto en df[col].
    Útil para:
      - Órdenes de fabricación: colorear por FECHA.
      - Orden de corte sugerida: colorear por MODELO.
    """
    if col not in df.columns:
        return df.style

    # valores únicos limpios
    uniques = []
    for v in df[col]:
        key = str(v).strip()
        if key and key not in uniques:
            uniques.append(key)

    if not uniques:
        return df.style

    # paleta ciclada (colores pastel imprimibles)
    palette = [
        "#f8bbd0",
        "#e1bee7",
        "#c5cae9",
        "#bbdefb",
        "#b2ebf2",
        "#b2dfdb",
        "#c8e6c9",
        "#dcedc8",
        "#fff9c4",
        "#ffe0b2",
    ]
    color_map = {u: palette[i % len(palette)] for i, u in enumerate(uniques)}

    def _styler(row):
        v = str(row.get(col, "")).strip()
        color = color_map.get(v, "")
        if not color:
            return pd.Series([""] * len(row), index=row.index)
        return _row_bg(row, color)

    return df.style.apply(lambda r: _styler(r), axis=1)


def _auto_qty_col(df, candidates=None):
    """
    Intenta encontrar la columna de cantidad en df.
    Devuelve nombre o None si no encuentra.
    """
    if candidates is None:
        candidates = [
            "STOCK",
            "stock",
            "STOCK_ESTIMADO",
            "stock_estimado",
            "ESTIMADO",
            "estimado",
            "QTY",
            "qty",
            "CANTIDAD",
            "cantidad",
            "TOTAL",
            "total",
        ]
    for c in candidates:
        if c in df.columns:
            return c
    # fallback: primer numérico que no parezca metadato
    blacklist = {
        "MODELO",
        "modelo",
        "TALLA",
        "talla",
        "PEDIDO",
        "pedido",
        "CLIENTE",
        "cliente",
        "COLOR",
        "color",
        "DESCRIPCION",
        "descripcion",
    }
    for col in df.columns:
        if col in blacklist:
            continue
        try:
            pd.to_numeric(df[col])
            return col
        except Exception:
            continue
    return None


from openpyxl.styles import Font, PatternFill


def _excel_yellow_header_and_total(ws, highlight_last: bool = True):
    """
    Cabeceras (fila 1) y, opcionalmente, última fila como TOTAL GENERAL,
    en amarillo chillón + negrita.
    """
    bright_fill = PatternFill(fill_type="solid", fgColor=BRIGHT_YELLOW)
    bold_font = Font(bold=True)

    # Cabecera
    for cell in ws[1]:
        cell.fill = bright_fill
        cell.font = bold_font

    # Total general (última fila) si aplica
    if highlight_last and ws.max_row >= 2:
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.fill = bright_fill
            cell.font = bold_font


def _excel_color_stock_ranges(ws, df: pd.DataFrame, qty_col_candidates=None):
    """
    Aplica semáforo por FILA en base a la columna de cantidad:
      <=0 rojo, 0-10 naranja, 10-25 amarillo pastel.
    """
    qty_col = _auto_qty_col(df, candidates=qty_col_candidates)
    if not qty_col:
        return

    try:
        col_idx = list(df.columns).index(qty_col) + 1  # 1-based
    except ValueError:
        return

    red_fill = PatternFill(fill_type="solid", fgColor="FFCDD2")
    orange_fill = PatternFill(fill_type="solid", fgColor="FFE0B2")
    yellow_fill = PatternFill(fill_type="solid", fgColor=PASTEL_YELLOW)

    # De la fila 2 hasta la última (totales los sobrescribiremos luego)
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        val = cell.value
        try:
            q = float(str(val).replace(",", "."))
        except Exception:
            continue

        fill = None
        if q <= 0:
            fill = red_fill
        elif 0 < q <= 10:
            fill = orange_fill
        elif 10 < q <= 25:
            fill = yellow_fill

        if fill:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=row, column=c).fill = fill


def _excel_color_pend_by_month(ws, df: pd.DataFrame, date_col: str = "FECHA"):
    """
    Mismas reglas que en la web para pedidos pendientes,
    pero aplicadas en el Excel (degradado verde↔rojo).
    """
    if date_col not in df.columns:
        return

    today = date.today()
    cur_idx = _month_index(today)
    date_col_idx = list(df.columns).index(date_col) + 1

    for row in range(2, ws.max_row):
        cell = ws.cell(row=row, column=date_col_idx)
        d = _parse_date_flexible(cell.value)
        if not d:
            continue
        idx = _month_index(d)
        delta = idx - cur_idx
        color = _month_delta_color(delta)
        if not color:
            continue

        fill = PatternFill(fill_type="solid", fgColor=color.replace("#", ""))
        for c in range(1, ws.max_column + 1):
            ws.cell(row=row, column=c).fill = fill


def _excel_color_by_column_palette(ws, df: pd.DataFrame, col: str):
    """
    Asigna un color por valor único en df[col] a todo el row en Excel.
    Usado para:
      - 04_ordenes_fabricacion → FECHA
      - 06_orden_corte_sugerida → MODELO
    """
    if col not in df.columns:
        return

    values = df[col].astype(str).str.strip().tolist()
    uniques = []
    for v in values:
        if v and v not in uniques:
            uniques.append(v)

    if not uniques:
        return

    palette = [
        "F8BBD0",
        "E1BEE7",
        "C5CAE9",
        "BBDEFB",
        "B2EBF2",
        "B2DFDB",
        "C8E6C9",
        "DCEDC8",
        "FFF9C4",
        "FFE0B2",
    ]
    color_map = {u: palette[i % len(palette)] for i, u in enumerate(uniques)}

    for row_idx, v in enumerate(values, start=2):
        color = color_map.get(v)
        if not color:
            continue
        fill = PatternFill(fill_type="solid", fgColor=color)
        for c in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=c).fill = fill


def _excel_highlight_totals_by_talla(ws, df: pd.DataFrame, talla_col: str = "TALLA"):
    """
    Pinta en amarillo chillón + negrita todas las filas donde TALLA sea:
      - 'TOTAL MODELO'
      - 'TOTAL GENERAL'
    """
    if talla_col not in df.columns:
        return

    try:
        talla_idx = list(df.columns).index(talla_col) + 1  # 1-based
    except ValueError:
        return

    bright_fill = PatternFill(fill_type="solid", fgColor=BRIGHT_YELLOW)
    bold_font = Font(bold=True)

    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=talla_idx).value
        if not val:
            continue
        txt = str(val).strip().upper()
        if txt in ("TOTAL MODELO", "TOTAL GENERAL"):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = bright_fill
                cell.font = bold_font


from openpyxl.styles import Border, Side


def _excel_add_borders(ws):
    """
    Añade borde fino (thin) a todas las celdas de la hoja.
    """
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = border


# --------------
# UI
# --------------
st.set_page_config(page_title="Gestor de Stock (GUI)", layout="wide")
# Cabecera con banner de "Última actualización"
col_title, col_status = st.columns([3, 2])
with col_title:
    st.title("🧰 Gestor de Stock y Previsión — UNIFORMIDAD - DEMO")
with col_status:
    lu = (
        get_last_update(st.session_state["manager"])
        if "manager" in st.session_state
        else None
    )
    html = """
    <div style="background:#f6f7fb;padding:10px 12px;border:1px solid #e6e6e6;
                border-radius:8px;font-size:14px;text-align:right;">
        <b>Última actualización:</b> {when} — {what}
    </div>
    """.format(
        when=(lu.get("timestamp_human") if lu else "—"),
        what=(lu.get("action") if lu else "—"),
    )
    st.markdown(html, unsafe_allow_html=True)


with st.sidebar:
    st.header("📂 Archivos (DEMO)")
    inv_path = st.text_input("Inventario JSON", DEFAULT_INV_PATH)
    prev_path = st.text_input("Previsión JSON", DEFAULT_PREV_PATH)
    tall_path = st.text_input("Talleres JSON", DEFAULT_TALL_PATH)
    cli_path = st.text_input("Clientes JSON", DEFAULT_CLI_PATH)

    if st.button("🔄 Cargar/Recargar"):
        # Invalida la cache del manager
        get_manager.clear()
        st.session_state["manager"] = get_manager(
            inv_path, prev_path, tall_path, cli_path
        )
        _success("Datos cargados.")
    if "manager" not in st.session_state:
        st.session_state["manager"] = get_manager(
            inv_path, prev_path, tall_path, cli_path
        )

    mgr: GestorStock = st.session_state["manager"]

    st.divider()
    st.caption(
        "Consejo: este MVP escribe en los mismos JSON. Haz copias si quieres probar sin riesgo."
    )

(
    tab_stock,
    tab_movs,
    tab_prevision,
    tab_auditoria,
    tab_catalogo,
    tab_imports,
    tab_backups,
    tab_export,
) = st.tabs(
    [
        "📦 Stock",
        "➡️ Movimientos",
        "📈 Previsión",
        "🧮 Auditoría",
        "📇 Catálogo & Maestros",
        "📥 Importaciones",
        "💾 Backups",
        "📤 Exportar CSV",
    ]
)

# -------------------
# TAB: STOCK
# -------------------
with tab_stock:
    st.subheader("Stock real")
    modelos = sorted(mgr.inventory.almacen.keys())
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        modelo_sel = st.selectbox("Modelo", ["(Todos)"] + modelos)
    with col2:
        talla_sel = st.text_input("Filtrar talla (opcional)").strip()
    with col3:
        st.write("")
        st.write("")
        if st.button("🔁 Refrescar listado", key="btn_stock_refresh"):
            pass  # la vista ya se reconstruye sola

    rows = []
    for m in modelos:
        if modelo_sel != "(Todos)" and m != modelo_sel:
            continue
        for t, q in sorted(
            mgr.inventory.almacen.get(m, {}).items(), key=lambda x: x[0]
        ):
            if talla_sel and norm_talla(t) != norm_talla(talla_sel):
                continue
            info = mgr.inventory.info_modelos.get(m, {})
            rows.append(
                {
                    "MODELO": m,
                    "DESCRIPCION": info.get("descripcion", ""),
                    "COLOR": info.get("color", ""),
                    "CLIENTE": info.get("cliente", ""),
                    "TALLA": t,
                    "STOCK": q,
                }
            )
    df = _to_df(rows)
    if not df.empty:
        st.dataframe(style_stock_ranges(df, qty_col="STOCK"), use_container_width=True)
    else:
        st.dataframe(df, use_container_width=True)

    with st.expander("📤 Exportar informes de Stock"):
        st.caption(
            f"Carpeta de exportación: `{getattr(mgr, 'EXPORT_DIR', '(no definida)')}`"
        )
        if st.button(
            "Exportar stock actual + (opcional) negativas", key="btn_export_stock"
        ):
            _run_export_all(mgr)

    with st.expander("🧹 Utilidades de saneo de stock"):
        col_u1, col_u2 = st.columns(2)

        # -- Ajustar NEGATIVOS a 0 --
        with col_u1:
            if st.button("Ajustar NEGATIVOS a 0", key="btn_fix_negatives"):
                try:
                    n, ruta_log, log_rows = _fix_negativos_a_cero_gui(mgr)
                    if n == 0:
                        _info("No había stock negativo.")
                    else:
                        _success(f"Ajustados {n} registros con stock negativo a 0.")
                        if log_rows:
                            st.dataframe(
                                pd.DataFrame(log_rows), use_container_width=True
                            )
                        if ruta_log:
                            st.caption(f"Log guardado en: `{ruta_log}`")
                        st.rerun()
                except Exception as e:
                    _error(f"Error ajustando negativos: {e}")

        # -- Reemplazar NaN/None/no enteros por 0 --
        with col_u2:
            if st.button("Reemplazar NaN/None/no enteros por 0", key="btn_fix_nans"):
                try:
                    n, log = _fix_bad_stock_values(mgr)
                    if n == 0:
                        _info("No había valores anómalos en el stock.")
                    else:
                        _success(f"Saneados {n} valores de stock.")
                        st.dataframe(pd.DataFrame(log), use_container_width=True)
                        st.rerun()
                except Exception as e:
                    _error(f"Error saneando stocks: {e}")

        # -- Purga de tallas anómalas --
        if st.button(
            "🧽 Eliminar tallas anómalas (NAN/NA/vacías) con stock 0",
            key="btn_purge_bad_tallas",
        ):
            try:
                n, ruta_log, log_rows = _purge_bad_talla_keys_gui(mgr, only_zero=True)
                if n == 0:
                    _info("No había tallas anómalas con stock 0 para eliminar.")
                else:
                    _success(f"Eliminadas {n} entradas de tallas anómalas.")
                    if log_rows:
                        st.dataframe(pd.DataFrame(log_rows), use_container_width=True)
                    if ruta_log:
                        st.caption(f"Log guardado en: `{ruta_log}`")
                    st.rerun()
            except Exception as e:
                _error(f"Error eliminando tallas anómalas: {e}")

    st.markdown("#### ✍️ Ajuste manual de stock")

    # --- Controles REACTIVOS (fuera del form) ---
    c1, c2, c3 = st.columns([2, 1, 1])

    with c1:
        labels, l2m = _modelo_labels_y_map(mgr)
        # si arriba en el listado has elegido un modelo, intenta preseleccionarlo
        pre_idx = 0
        if modelo_sel and modelo_sel != "(Todos)":
            for i, lab in enumerate(labels, start=1):
                if lab.startswith(modelo_sel):
                    pre_idx = i
                    break
        sel_label = st.selectbox(
            "Modelo", options=[""] + labels, index=pre_idx, key="ajuste_modelo_lbl"
        )
        m_m = l2m.get(sel_label, "").upper().strip()

    with c2:
        # selector de talla con opciones conocidas del modelo (y opción manual)
        m_t = talla_select(
            "Talla", m_m, key_sel="ajuste_talla_sel", key_txt="ajuste_talla_txt"
        )

    # calcula el stock actual con el helper GLOBAL _stock_actual (definido arriba)
    actual = _stock_actual(mgr, m_m, m_t)

    with c3:
        st.number_input(
            "Stock actual",
            min_value=0,
            step=1,
            value=int(actual),
            key="ajuste_stock_actual_view",
            disabled=True,
        )

    # --- Form SOLO para confirmar/escribir ---
    with st.form("form_ajuste_manual"):
        f1, f2 = st.columns([1, 3])
        with f1:
            # por defecto proponemos el stock actual
            nuevo = st.number_input(
                "Nuevo stock",
                min_value=0,
                step=1,
                value=int(actual),
                key="ajuste_nuevo_stock",
            )
        with f2:
            obs = st.text_input(
                "Observación",
                value=f"Ajuste manual via GUI {_timestamp()}",
                key="ajuste_obs",
            )

        sub_am = st.form_submit_button("Aplicar ajuste")

        if sub_am:
            if not m_m or not m_t:
                _error("Modelo y talla son obligatorios.")
            else:
                try:
                    delta = int(nuevo) - int(actual)
                    if delta == 0:
                        _warn("No hay cambios que aplicar.")
                    else:
                        cambios = [
                            {
                                "modelo": m_m,
                                "talla": norm_talla(m_t),
                                "antes": int(actual),
                                "despues": int(nuevo),
                                "delta": int(nuevo) - int(actual),
                                "observacion": obs,
                            }
                        ]
                        n = mgr.inventory.apply_stock_fixes(cambios)

                        _success(f"Ajuste aplicado. Registros modificados: {n}")
                        set_last_update(
                            mgr,
                            f"Ajuste stock {m_m} T:{m_t} Δ{int(nuevo) - int(actual)}",
                        )
                        st.rerun()  # refresca para que “Stock actual” se actualice al instante

                except Exception as e:
                    _error(f"Error aplicando ajuste: {e}")


# -------------------
# TAB: MOVIMIENTOS
# -------------------
with tab_movs:
    st.subheader("Registrar entradas/salidas")

    tab_e, tab_s = st.tabs(["➕ Entrada", "➖ Salida"])

    # ENTRADA
    with tab_e:
        # Modelo fuera del form para reactividad de tallas
        mcol1, mcol2 = st.columns(2)
        with mcol1:
            labels_in, l2m_in = _modelo_labels_y_map(mgr)
            sel_label_in = st.selectbox(
                "Modelo", options=[""] + labels_in, key="entrada_modelo_lbl"
            )
            modelo = l2m_in.get(sel_label_in, "").upper().strip()
        with mcol2:
            fecha = st.text_input(
                "Fecha (YYYY-MM-DD)",
                value=datetime.now().strftime("%Y-%m-%d"),
                key="ent_fecha",
            )

        talleres = [t.nombre for t in mgr.workshops.list_all()]
        tcol1, tcol2, tcol3 = st.columns([2, 1, 1])
        with tcol1:
            taller = st.selectbox("Taller (opcional)", [""] + talleres)
        with tcol2:
            talla = talla_select(
                "Talla",
                modelo,
                key_sel="entrada_talla_sel",
                key_txt="entrada_talla_txt",
            )
        with tcol3:
            cantidad = st.number_input("Cantidad", min_value=1, step=1, value=1)

        obs = st.text_input("Observaciones (opcional)", value="")

        if st.button("Guardar entrada", key="btn_guardar_entrada"):
            if not modelo or not talla:
                _error("Modelo y talla son obligatorios.")
            else:
                try:
                    mgr.inventory.register_entry(
                        modelo=modelo,
                        talla=norm_talla(talla),
                        cantidad=int(cantidad),
                        taller=taller,
                        fecha=fecha or None,
                        proveedor="",
                        observaciones=obs,
                    )
                    _success("Entrada registrada.")
                    set_last_update(mgr, f"Entrada {modelo} T:{talla} +{int(cantidad)}")
                    st.rerun()

                except Exception as e:
                    _error(f"Error registrando entrada: {e}")

    # SALIDA
    with tab_s:
        # Modelo fuera del form para reactividad de tallas
        mcol1, mcol2 = st.columns(2)
        with mcol1:
            labels_out, l2m_out = _modelo_labels_y_map(mgr)
            sel_label_out = st.selectbox(
                "Modelo", options=[""] + labels_out, key="salida_modelo_lbl"
            )
            modelo = l2m_out.get(sel_label_out, "").upper().strip()
        with mcol2:
            fecha = st.text_input(
                "Fecha (YYYY-MM-DD)",
                value=datetime.now().strftime("%Y-%m-%d"),
                key="s_fecha",
            )

        clientes = [c.nombre for c in mgr.clients.list_all()]
        scol1, scol2, scol3, scol4 = st.columns([2, 1, 1, 1])
        with scol1:
            cliente = st.selectbox(
                "Cliente (opcional)", [""] + clientes, key="s_cliente"
            )
        with scol2:
            pedido = st.text_input("Pedido", key="s_pedido")
        with scol3:
            albaran = st.text_input("Albarán", key="s_albaran")
        with scol4:
            talla = talla_select(
                "Talla", modelo, key_sel="salida_talla_sel", key_txt="salida_talla_txt"
            )
        cant = st.number_input(
            "Cantidad", min_value=1, step=1, value=1, key="s_cantidad"
        )

        if st.button("Guardar salida", key="btn_guardar_salida"):
            if not modelo or not talla or not pedido or not albaran:
                _error("Modelo, talla, pedido y albarán son obligatorios.")
            else:
                try:
                    ok = mgr.inventory.register_exit(
                        modelo=modelo,
                        talla=norm_talla(talla),
                        cantidad=int(cant),
                        cliente=cliente,
                        pedido=norm_codigo(pedido),
                        albaran=norm_codigo(albaran),
                        fecha=fecha or None,
                    )
                    _success(
                        "Salida registrada."
                        if ok
                        else "No se pudo registrar la salida."
                    )
                    if ok:
                        set_last_update(
                            mgr,
                            f"Salida {modelo} T:{talla} -{int(cant)} Ped:{pedido} Alb:{albaran}",
                        )
                    st.rerun()

                except Exception as e:
                    _error(f"Error registrando salida: {e}")

    st.divider()
    with st.expander("📤 Exportar informes de Movimientos"):
        st.caption(
            f"Carpeta de exportación: `{getattr(mgr, 'EXPORT_DIR', '(no definida)')}`"
        )
        if st.button("Exportar entradas/salidas (pack)", key="btn_export_movs"):
            _run_export_all(mgr)
# -------------------
# TAB: PREVISIÓN
# -------------------
with tab_prevision:
    st.subheader("Stock estimado (Real + Fabricación - Pendientes)")

    colf1, colf2 = st.columns([1, 3])
    with colf1:
        if st.button("🔁 Recalcular", key="btn_prev_recalc"):
            st.rerun()
    est = mgr.prevision.calc_estimated_stock(mgr.inventory)
    est_df = pd.DataFrame(est).sort_values(["modelo", "talla"])
    if not est_df.empty:
        st.dataframe(
            style_stock_ranges(est_df, qty_col="stock_estimado"),
            use_container_width=True,
        )
    else:
        st.dataframe(est_df, use_container_width=True)

    st.divider()
    st.markdown("### Pedidos pendientes")
    # Listado simple
    pend = mgr.prevision.list_pendings()
    pend_rows = []
    for idx, p in pend:
        info = mgr.inventory.info_modelos.get(
            p["modelo"], mgr.prevision.info_modelos.get(p["modelo"], {})
        )
        pend_rows.append(
            {
                "IDX": idx,
                "MODELO": p["modelo"],
                "DESCRIPCION": info.get("descripcion", ""),
                "COLOR": info.get("color", ""),
                "TALLA": p.get("talla", ""),
                "CANTIDAD": p.get("cantidad", 0),
                "PEDIDO": p.get("pedido", ""),
                "NUMERO_PEDIDO": p.get("numero_pedido", ""),
                "CLIENTE": p.get("cliente", ""),
                "FECHA": p.get("fecha", ""),
            }
        )

    cols = [
        "IDX",
        "MODELO",
        "DESCRIPCION",
        "COLOR",
        "TALLA",
        "CANTIDAD",
        "PEDIDO",
        "NUMERO_PEDIDO",
        "CLIENTE",
        "FECHA",
    ]
    if pend_rows:
        df_pend = _to_df(pend_rows)
        view_cols = [c for c in cols if c in df_pend.columns]

        if not df_pend.empty and "FECHA" in df_pend.columns:
            st.dataframe(
                style_pend_by_month(df_pend[view_cols], date_col="FECHA"),
                use_container_width=True,
            )
        else:
            st.dataframe(df_pend[view_cols], use_container_width=True)

    else:
        st.info("No hay pedidos pendientes.")

    with st.expander("➕ Añadir pendiente"):
        # Modelo fuera del form para reactividad de tallas
        pcol1, pcol2, pcol3, pcol4 = st.columns([2, 1, 1, 1])
        with pcol1:
            labels_p, l2m_p = _modelo_labels_y_map(mgr)
            sel_label_p = st.selectbox("Modelo", options=labels_p, key="pend_m_lbl")
            p_modelo = l2m_p.get(sel_label_p, "").upper().strip()
        with pcol2:
            p_talla = talla_select(
                "Talla", p_modelo, key_sel="pend_talla_sel", key_txt="pend_talla_txt"
            )
        with pcol3:
            p_cant = st.number_input(
                "Cantidad", min_value=1, step=1, value=1, key="pend_c"
            )
        with pcol4:
            p_fecha = st.text_input(
                "Fecha (YYYY-MM-DD)",
                value=datetime.now().strftime("%Y-%m-%d"),
                key="pend_f",
            )
        clientes = [c.nombre for c in mgr.clients.list_all()]
        p_cliente = st.selectbox("Cliente", options=clientes, key="pend_cli")
        p_pedido = st.text_input("Pedido", key="pend_ped")
        p_num_int = st.text_input("Número interno (opcional)", key="pend_numint")

        if st.button("Añadir", key="btn_pend_anadir"):
            try:
                mgr.prevision.register_pending(
                    modelo=p_modelo,
                    talla=norm_talla(p_talla),
                    cantidad=int(p_cant),
                    pedido=norm_codigo(p_pedido),
                    cliente=p_cliente,
                    fecha=p_fecha or None,
                    numero_pedido=norm_codigo(p_num_int) or None,
                )
                _success("Pedido pendiente añadido.")
                set_last_update(
                    mgr,
                    f"Pendiente + {p_modelo} T:{p_talla} Q:{int(p_cant)} Ped:{p_pedido}",
                )
                st.rerun()
            except Exception as e:
                _error(f"Error: {e}")

    st.divider()
    with st.expander("✏️ Editar / 🗑️ Eliminar pedidos pendientes"):
        cedit, cdel = st.columns(2)

        # -------- EDITAR PENDIENTE --------
        with cedit:
            st.markdown("**Editar por IDX**")
            with st.form("form_edit_pending"):
                if pend_rows:
                    sel_pend_edit = st.selectbox(
                        "Selecciona pedido a editar",
                        options=pend_rows,  # pasa los dicts
                        format_func=_fmt_pending_label,  # cómo mostrarlos
                        key="pend_sel_edit",
                    )
                    idx_ed = sel_pend_edit["IDX"]
                else:
                    st.warning("No hay pedidos para editar.")
                    idx_ed = None

                ecol1, ecol2 = st.columns(2)
                with ecol1:
                    e_modelo = (
                        st.text_input(
                            "Modelo (vacío = sin cambio)", key="pend_edit_modelo"
                        )
                        .upper()
                        .strip()
                    )
                    e_talla = st.text_input(
                        "Talla (vacío = sin cambio)", key="pend_edit_talla"
                    )
                    e_pedido = st.text_input(
                        "Pedido (vacío = sin cambio)", key="pend_edit_pedido"
                    )
                    e_num = st.text_input(
                        "Número interno (vacío = sin cambio)", key="pend_edit_num"
                    )
                with ecol2:
                    e_cliente = st.text_input(
                        "Cliente (vacío = sin cambio)", key="pend_edit_cliente"
                    )
                    e_fecha = st.text_input(
                        "Fecha YYYY-MM-DD (vacío = sin cambio)", key="pend_edit_fecha"
                    )
                    e_cant_str = st.text_input(
                        "Cantidad (vacío = sin cambio)", key="pend_edit_cant"
                    )

                sub_edit = st.form_submit_button("Aplicar cambios")
                if sub_edit:
                    if idx_ed is None:
                        _warn("No hay pedidos para editar.")
                    else:
                        try:
                            e_cant = int(e_cant_str) if e_cant_str.strip() else None
                            mgr.prevision.edit_pending(
                                int(idx_ed),
                                modelo=e_modelo or None,
                                talla=norm_talla(e_talla) if e_talla.strip() else None,
                                cantidad=e_cant,
                                pedido=(
                                    norm_codigo(e_pedido) if e_pedido.strip() else None
                                ),
                                cliente=e_cliente or None,
                                fecha=e_fecha or None,
                                numero_pedido=(
                                    norm_codigo(e_num) if e_num.strip() else None
                                ),
                            )
                            _success("Pedido pendiente actualizado.")
                            set_last_update(mgr, f"Editar pendiente IDX:{idx_ed}")
                            st.rerun()
                        except Exception as e:
                            _error(f"Error: {e}")

        # -------- ELIMINAR PENDIENTE --------
        with cdel:
            st.markdown("**Eliminar por IDX**")
            with st.form("form_del_pending"):
                if pend_rows:
                    sel_pend_del = st.selectbox(
                        "Selecciona pedido a eliminar",
                        options=pend_rows,
                        format_func=_fmt_pending_label,
                        key="pend_sel_del",
                    )
                    idx_del = sel_pend_del["IDX"]
                else:
                    st.warning("No hay pedidos para eliminar.")
                    idx_del = None

                sub_del = st.form_submit_button("Eliminar")
                if sub_del:
                    if idx_del is None:
                        _warn("No hay pedidos para eliminar.")
                    else:
                        try:
                            mgr.prevision.delete_pending(int(idx_del))
                            _success("Pedido pendiente eliminado.")
                            set_last_update(mgr, f"Eliminar pendiente IDX:{idx_del}")
                            st.rerun()
                        except Exception as e:
                            _error(f"Error: {e}")

    st.divider()
    st.markdown("### Órdenes de fabricación")
    items = mgr.prevision.list_fabrication()
    fab_rows = []
    for idx, it in items:
        info = mgr.inventory.info_modelos.get(
            it["modelo"], mgr.prevision.info_modelos.get(it["modelo"], {})
        )
        fab_rows.append(
            {
                "IDX": idx,
                "MODELO": it["modelo"],
                "DESCRIPCION": info.get("descripcion", ""),
                "COLOR": info.get("color", ""),
                "TALLA": it.get("talla", ""),
                "CANTIDAD": it.get("cantidad", 0),
                "FECHA": it.get("fecha", ""),
            }
        )

    cols_fab = ["IDX", "MODELO", "DESCRIPCION", "COLOR", "TALLA", "CANTIDAD", "FECHA"]
    if fab_rows:
        df_fab = _to_df(fab_rows)
        view = df_fab[[c for c in cols_fab if c in df_fab.columns]]
        if "FECHA" in view.columns:
            st.dataframe(
                style_by_column_palette(view, "FECHA"), use_container_width=True
            )
        else:
            st.dataframe(view, use_container_width=True)
    else:
        st.info("No hay órdenes de fabricación.")

    c1, c2 = st.columns(2)
    with c1:
        # Modelo fuera del form para reactividad de tallas
        fcol1, fcol2, fcol3 = st.columns([2, 1, 1])
        with fcol1:
            labels_f, l2m_f = _modelo_labels_y_map(mgr)
            sel_label_f = st.selectbox("Modelo", options=labels_f, key="fab_m_lbl")
            f_modelo = l2m_f.get(sel_label_f, "").upper().strip()
        with fcol2:
            f_talla = talla_select(
                "Talla", f_modelo, key_sel="fab_talla_sel", key_txt="fab_talla_txt"
            )
        with fcol3:
            f_cant = st.number_input(
                "Cantidad", min_value=1, step=1, value=1, key="fab_c"
            )
        f_fecha = st.text_input(
            "Fecha (YYYY-MM-DD)", value=datetime.now().strftime("%Y-%m-%d"), key="fab_f"
        )

        if st.button("Añadir orden", key="btn_fab_anadir"):
            try:
                mgr.prevision.register_order(
                    f_modelo, norm_talla(f_talla), int(f_cant), fecha=f_fecha or None
                )
                _success("Orden de fabricación añadida.")
                set_last_update(
                    mgr, f"Orden fabricación + {f_modelo} T:{f_talla} Q:{int(f_cant)}"
                )
                st.rerun()
            except Exception as e:
                _error(f"Error: {e}")

    with c2:
        with st.form("form_edit_del_fab"):
            if fab_rows:
                opciones_fab = [
                    f"{f['IDX']} | {f['MODELO']} | {f['TALLA']} | {f['FECHA']}"
                    for f in fab_rows
                ]
                sel_fab_edit = st.selectbox(
                    "Selecciona orden a editar/eliminar",
                    opciones_fab,
                    key="fab_sel_edit",
                )
                idx_edit = int(sel_fab_edit.split("|")[0].strip())
            else:
                st.warning("No hay órdenes de fabricación.")
                idx_edit = None

            nueva = st.number_input(
                "Nueva cantidad (0=eliminar)", min_value=0, step=1, value=0
            )
            subed = st.form_submit_button("Aplicar cambio")
            if subed:
                if idx_edit is None:
                    _warn("No hay órdenes de fabricación.")
                else:
                    try:
                        mgr.prevision.edit_fabrication_qty(int(idx_edit), int(nueva))
                        _success("Orden actualizada/eliminada.")
                        set_last_update(
                            mgr,
                            f"Editar/Eliminar orden fabricación IDX:{idx_edit} Nueva:{int(nueva)}",
                        )
                        st.rerun()
                    except Exception as e:
                        _error(f"Error: {e}")

    st.divider()
    with st.expander("📤 Exportar informes de Previsión"):
        st.caption(
            f"Carpeta de exportación: `{getattr(mgr, 'EXPORT_DIR', '(no definida)')}`"
        )
        if st.button(
            "Exportar pendientes/órdenes/estimado (pack)", key="btn_export_prev"
        ):
            _run_export_all(mgr)
# -------------------
# TAB: AUDITORÍA
# -------------------
with tab_auditoria:
    st.subheader("Auditoría de stock vs histórico")
    solo_modelo = (
        st.text_input("Filtrar por modelo (opcional)", value="").upper().strip() or None
    )
    if st.button("🔎 Auditar", key="btn_audit_go"):
        pass

    cambios = mgr.inventory.audit_and_fix_stock(aplicar=False, solo_modelo=solo_modelo)
    if not cambios:
        _success("Sin desajustes. Todo cuadra.")
    else:
        st.write(f"Encontradas **{len(cambios)}** diferencias.")
        dfc = pd.DataFrame(cambios)
        dfc.index = dfc.index + 1  # que se vea 1..N
        st.dataframe(dfc, use_container_width=True)

        st.markdown("#### Aplicar ajustes (modifican el stock real)")
        opciones = {
            "Todos": cambios,
            "Solo Δ positivo (sube stock)": [r for r in cambios if r["delta"] > 0],
            "Solo Δ negativo (baja stock)": [r for r in cambios if r["delta"] < 0],
        }
        modo = st.selectbox("Modo", list(opciones.keys()))
        aplicar = opciones[modo]

        idx_text = st.text_input("Índices concretos (ej. 1,3,5-8) — opcional")
        if idx_text.strip():
            idxs = parse_index_selection(idx_text, len(cambios))

            if idxs:
                aplicar = [cambios[i - 1] for i in idxs]
            else:
                _warn(
                    "Selección vacía o fuera de rango: se ignora y se aplicará la opción de arriba."
                )

        if st.button(
            f"🛠️ Aplicar {len(aplicar)} ajustes de stock", key="btn_audit_apply"
        ):
            try:
                n = mgr.inventory.apply_stock_fixes(aplicar)
                _success(f"Ajustes aplicados: {n}")
                set_last_update(mgr, f"Auditoría: {n} ajustes aplicados")
            except Exception as e:
                _error(f"Error aplicando ajustes: {e}")

        st.divider()
        st.markdown("#### Regularizar HISTÓRICO (no toca stock real)")
        obs = st.text_input(
            "Observación", value="Ajuste auditoría (GUI)", key="audit_reg_obs"
        )
        fecha = st.text_input(
            "Fecha (YYYY-MM-DD)",
            value=datetime.now().strftime("%Y-%m-%d"),
            key="audit_reg_fecha",
        )

        opciones2 = {
            "Todos": cambios,
            "Solo Δ positivo → SALIDAS de ajuste": [
                r for r in cambios if r["delta"] > 0
            ],
            "Solo Δ negativo → ENTRADAS de ajuste": [
                r for r in cambios if r["delta"] < 0
            ],
        }
        modo2 = st.selectbox("Modo regularización", list(opciones2.keys()))
        aplicar2 = opciones2[modo2]

        idx_text2 = st.text_input(
            "Índices concretos para regularizar (opcional)", key="idx2"
        )
        if idx_text2.strip():
            idxs2 = parse_index_selection(idx_text2, len(cambios))
            if idxs2:
                aplicar2 = [cambios[i - 1] for i in idxs2]
            else:
                _warn("Selección vacía o fuera de rango: se ignora.")

        if st.button(
            f"🧾 Crear {len(aplicar2)} asientos de regularización",
            key="btn_audit_regularize",
        ):
            try:
                n = mgr.inventory.regularize_history_to_current(
                    aplicar2, fecha=fecha, observacion=obs
                )
                _success(f"Asientos creados: {n}")
                set_last_update(mgr, f"Regularización histórica: {n} asientos")
            except Exception as e:
                _error(f"Error creando asientos: {e}")

    st.divider()
    with st.expander("📤 Exportar informes de Auditoría"):
        st.caption(
            f"Carpeta de exportación: `{getattr(mgr, 'EXPORT_DIR', '(no definida)')}`"
        )
        cols = st.columns(2)
        with cols[0]:
            if st.button("Exportar auditoría/paquete completo"):
                _run_export_all(mgr)
        with cols[1]:
            if st.button("Solo stock NEGATIVO", key="btn_negativos_auditoria"):
                _run_export_stock_negativo(mgr)
# -------------------
# TAB: CATÁLOGO & MAESTROS
# -------------------
with tab_catalogo:
    st.subheader("Modelos (catálogo)")
    rows = []
    for m, info in sorted(mgr.inventory.info_modelos.items()):
        rows.append({"MODELO": m, **info})
    st.dataframe(_to_df(rows), use_container_width=True)

    st.markdown("#### Editar info de modelo")
    with st.form("form_model_info"):
        mi1, mi2, mi3, mi4 = st.columns([2, 2, 2, 1])
        with mi1:
            m_m = st.text_input("Modelo", key="mi_m").upper().strip()
        with mi2:
            m_desc = st.text_input("Descripción (opcional)", key="mi_d")
        with mi3:
            m_color = st.text_input("Color (opcional)", key="mi_c")
        with mi4:
            m_cli = st.text_input("Cliente (opcional)", key="mi_cli")
        sub_mi = st.form_submit_button("Guardar")
        if sub_mi:
            if m_m:
                try:
                    mgr.inventory.update_model_info(
                        modelo=m_m,
                        descripcion=m_desc or None,
                        color=m_color or None,
                        cliente=m_cli or None,
                    )
                    _success("Modelo actualizado.")
                except Exception as e:
                    _error(f"Error: {e}")
            else:
                _error("Indica modelo.")

    st.divider()
    st.subheader("Talleres")
    t_rows = [
        {"NOMBRE": t.nombre, "CONTACTO": t.contacto or ""}
        for t in mgr.workshops.list_all()
    ]
    st.dataframe(_to_df(t_rows), use_container_width=True)

    with st.form("form_add_taller"):
        t1, t2 = st.columns([2, 2])
        with t1:
            t_nombre = st.text_input("Nombre taller").strip()
        with t2:
            t_contacto = st.text_input("Contacto (opcional)").strip()
        sub_t = st.form_submit_button("Añadir taller")
        if sub_t:
            if not t_nombre:
                _error("Nombre obligatorio.")
            else:
                try:
                    mgr.workshops.add(t_nombre, t_contacto or None)
                    _success("Taller añadido.")
                except Exception as e:
                    _error(f"Error: {e}")

    st.divider()
    st.subheader("Clientes")
    c_rows = [
        {"NOMBRE": c.nombre, "CONTACTO": c.contacto or ""}
        for c in mgr.clients.list_all()
    ]
    st.dataframe(_to_df(c_rows), use_container_width=True)

    with st.form("form_add_cliente"):
        c1, c2 = st.columns([2, 2])
        with c1:
            c_nombre = st.text_input("Nombre cliente").strip()
        with c2:
            c_contacto = st.text_input("Contacto (opcional)").strip()
        sub_c = st.form_submit_button("Añadir cliente")
        if sub_c:
            if not c_nombre:
                _error("Nombre obligatorio.")
            else:
                try:
                    mgr.clients.add(c_nombre, c_contacto or None)
                    _success("Cliente añadido.")
                except Exception as e:
                    _error(f"Error: {e}")

# -------------------
# TAB: IMPORTACIONES
# -------------------
with tab_imports:
    st.subheader("📥 Importaciones")

    # ---------- ALBARANES SERVIDOS ----------
    st.markdown("### 🚚 Importar albaranes servidos (Excel)")

    col_alb_left, col_alb_right = st.columns(2)

    # Opción 1: subir archivo
    with col_alb_left:
        st.markdown("**Subir el Excel de albaranes**")
        up_alb = st.file_uploader(
            "Arrastra o selecciona Excel", type=["xlsx", "xls"], key="alb_upl"
        )
        modo_dup = st.selectbox(
            "Líneas duplicadas (mismo MODELO/TALLA/PEDIDO/ALBARÁN ya registradas)",
            [
                "Descontar diferencia (recomendado)",
                "Ignorar duplicadas",
                "Procesar todo igualmente",
            ],
            key="alb_dup_upl",
        )
        skip_alb = st.number_input(
            "Filas a saltar (header)", min_value=0, step=1, value=25, key="alb_skip_upl"
        )
        simular_alb = st.checkbox(
            "Simular (no escribir)", value=False, key="alb_sim_upl"
        )
        if st.button("Procesar albaranes (archivo subido)", key="btn_alb_upl"):
            if not up_alb:
                _error("Sube un Excel primero.")
            else:
                df = pd.read_excel(io.BytesIO(up_alb.read()), skiprows=int(skip_alb))
                _procesar_albaranes_df(df, modo_dup, simular_alb)

    # Opción 2: ruta fija (la de tu CLI)
    with col_alb_right:
        st.markdown("**Usar ruta fija (igual que el script)**")
        st.caption(
            f"Ruta configurada en el gestor: `{getattr(mgr, 'ALBARANES_EXCEL', '(no definida)')}`"
        )
        modo_dup_fx = st.selectbox(
            "Líneas duplicadas",
            [
                "Descontar diferencia (recomendado)",
                "Ignorar duplicadas",
                "Procesar todo igualmente",
            ],
            key="alb_dup_fx",
        )
        simular_alb_fx = st.checkbox(
            "Simular (no escribir)", value=False, key="alb_sim_fx"
        )
        if st.button("Procesar albaranes (ruta fija)", key="btn_alb_fx"):
            ruta = getattr(mgr, "ALBARANES_EXCEL", None)
            if not ruta:
                _error("No hay ruta fija configurada en el gestor (ALBARANES_EXCEL).")
            else:
                df = pd.read_excel(ruta, skiprows=25)  # como en la versión CLI
                _procesar_albaranes_df(df, modo_dup_fx, simular_alb_fx)

    st.divider()

    # ---------- PEDIDOS PENDIENTES ----------
    st.markdown("### 🧾 Importar pedidos pendientes (Excel)")

    col_ped_left, col_ped_right = st.columns(2)

    # Opción 1: subir archivo
    with col_ped_left:
        st.markdown("**Subir el Excel de pedidos**")
        up_ped = st.file_uploader(
            "Arrastra o selecciona Excel", type=["xlsx", "xls"], key="ped_upl"
        )
        skip_ped = st.number_input(
            "Filas a saltar (header)", min_value=0, step=1, value=26, key="ped_skip_upl"
        )
        simular_ped = st.checkbox(
            "Simular (no escribir)", value=False, key="ped_sim_upl"
        )
        if st.button("Procesar pedidos (archivo subido)", key="btn_ped_upl"):
            if not up_ped:
                _error("Sube un Excel primero.")
            else:
                df = pd.read_excel(io.BytesIO(up_ped.read()), skiprows=int(skip_ped))
                _procesar_pedidos_df(df, simular_ped)

    # Opción 2: ruta fija (la de tu CLI)
    with col_ped_right:
        st.markdown("**Usar ruta fija (igual que el script)**")
        st.caption(
            f"Ruta configurada en el gestor: `{getattr(mgr, 'PEDIDOS_EXCEL', '(no definida)')}`"
        )
        simular_ped_fx = st.checkbox(
            "Simular (no escribir)", value=False, key="ped_sim_fx"
        )
        if st.button("Procesar pedidos (ruta fija)", key="btn_ped_fx"):
            ruta = getattr(mgr, "PEDIDOS_EXCEL", None)
            if not ruta:
                _error("No hay ruta fija configurada en el gestor (PEDIDOS_EXCEL).")
            else:
                df = pd.read_excel(ruta, skiprows=26)  # como en la versión CLI
                _procesar_pedidos_df(df, simular_ped_fx)


# -------------------
# TAB: EXPORTAR CSV
# -------------------
with tab_export:
    st.subheader("📤 Exportar CSV / Excel (pack completo o selectivo)")

    # Mostrar rutas de exportación de forma clara
    export_dir = getattr(mgr, "EXPORT_DIR", "(no definida)")
    export_csv_dir = getattr(mgr, "EXPORT_CSV_DIR", "(no definida)")
    export_imprimir_dir = getattr(mgr, "IMPRIMIR_DIR", "(no definida)")

    st.caption(f"📁 Carpeta raíz de exportaciones: `{export_dir}`")
    st.caption(f"📄 CSV: `{export_csv_dir}`")
    st.caption(f"📊 Excels IMPRIMIR: `{export_imprimir_dir}`")

    cols = st.columns(3)
    with cols[0]:
        if st.button("Exportar TODO (pack completo CSV)", key="btn_export_all"):
            _run_export_all(mgr)
    with cols[1]:
        if st.button("Solo stock NEGATIVO", key="btn_negativos_export"):
            _run_export_stock_negativo(mgr)
    with cols[2]:
        if st.button("Recalcular y exportar de nuevo CSV", key="btn_export_recalc_all"):
            _run_export_all(mgr)

    st.divider()
    st.markdown("### 🧾 Exportar Excel listos para imprimir (IMPRIMIR_XX_...)")
    if st.button(
        "Generar Excel IMPRIMIR (00,03,04,05,06)", key="btn_export_excel_pack"
    ):
        _export_excel_pack(mgr)


# -------------------
# TAB: BACKUPS
# -------------------
with tab_backups:
    st.subheader("💾 Copias de seguridad")
    back_dir = os.path.join(os.path.dirname(mgr.ds_inventario.path), "backups")
    os.makedirs(back_dir, exist_ok=True)
    st.caption(f"Carpeta de backups: `{back_dir}`")

    col_b1, col_b2 = st.columns(2)
    with col_b1:
        if st.button("Crear backup ahora", key="btn_backup_create_main"):
            try:
                fecha = _timestamp()
                ruta_datos = os.path.join(back_dir, f"datos_almacen_{fecha}.json")
                ruta_prevision = os.path.join(back_dir, f"prevision_{fecha}.json")
                with open(mgr.ds_inventario.path, "r", encoding="utf-8") as src, open(
                    ruta_datos, "w", encoding="utf-8"
                ) as dst:
                    dst.write(src.read())
                with open(mgr.ds_prevision.path, "r", encoding="utf-8") as src, open(
                    ruta_prevision, "w", encoding="utf-8"
                ) as dst:
                    dst.write(src.read())
                _success(f"Backup creado:\n- {ruta_datos}\n- {ruta_prevision}")
            except Exception as e:
                _error(f"Error creando backup: {e}")

    with col_b2:
        archivos = [f for f in os.listdir(back_dir) if f.endswith(".json")]
        archivos.sort(reverse=True)
        sel = st.selectbox("Selecciona backup a restaurar", ["(ninguno)"] + archivos)
        if st.button("Restaurar seleccionado", key="btn_backup_restore_main"):
            if sel == "(ninguno)":
                _warn("Elige un archivo de backup.")
            else:
                try:
                    origen = os.path.join(back_dir, sel)
                    if "datos_almacen" in sel:
                        destino = mgr.ds_inventario.path
                    elif "prevision" in sel:
                        destino = mgr.ds_prevision.path
                    else:
                        _error(
                            "Nombre de backup no reconocido (debe incluir 'datos_almacen' o 'prevision')."
                        )
                        destino = None
                    if destino:
                        with open(origen, "r", encoding="utf-8") as src, open(
                            destino, "w", encoding="utf-8"
                        ) as dst:
                            dst.write(src.read())
                        get_manager.clear()
                        st.session_state["manager"] = get_manager(
                            inv_path, prev_path, tall_path, cli_path
                        )
                        _success(f"Restaurado '{sel}' en {destino}")
                        set_last_update(mgr, f"Restaurado backup: {sel}")
                        st.rerun()
                except Exception as e:
                    _error(f"Error restaurando backup: {e}")


st.caption(
    "MVP+ Streamlit • Usa las mismas rutas JSON del script original • Import/Backup incluidos • by Aitor Susperregui"
)
